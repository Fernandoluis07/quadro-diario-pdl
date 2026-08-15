import datetime
import hashlib
import json
import subprocess

import openpyxl
import pandas as pd
import pytest

from backend import cabecalho, config
from backend.tests.test_congelar import INDEX_HTML_SINTETICO


def _linha_mb51(material, deposito, bwart, data, referencia, valor=10.0, ordem=None, centro_custo=None):
    return {
        "Material": material, "Texto breve material": "x", "Centro": "2003",
        "Depósito": deposito, "Tipo de movimento": bwart, "Data de lançamento": data,
        "Qtd.  UM registro": 1, "UM registro": "UN", "Montante em MI": valor,
        "Reserva": "R1", "Referência": referencia, "Nome do usuário": "fernando",
        "Ordem": ordem, "Centro custo": centro_custo,
    }


def _preparar_bases(tmp_path):
    bases_dir = tmp_path / "Bases"
    bases_dir.mkdir()

    mb51 = pd.DataFrame(
        [
            _linha_mb51("1001", "D009", 201, "09.08.2026", "NF-0", valor=5.0),
            _linha_mb51("1002", "D009", 201, "10.08.2026", "NF-1", valor=10.0),
            _linha_mb51("1003", "D009", 101, "10.08.2026", "NF-2", valor=4.0),
        ]
    )
    mb51.to_excel(bases_dir / config.MB51_FILENAME, index=False)

    mb25 = pd.DataFrame(
        [{"Reserva": "R10", "Material": "2001", "Texto breve material": "Item A", "Depósito": "D009",
          "Qtd.necessária": 3, "Data da necessidade": "15/08/2026", "Centro custo": None, "Ordem": "O10"}]
    )
    mb25.to_excel(bases_dir / config.MB25_FILENAME, index=False)

    zmm028 = pd.DataFrame(
        [
            {"Material": "3001", "Denom.": "A", "Unidade": "UN", "Centro": "2003", "Util.livre": 10,
             "Val.total": 100.0, "Pos.dpst.": "A-01", "Tp.MRP": "VB", "Depósito": "D009",
             "Estq.máx.": 50, "Pt.reabast": 10},
        ]
    )
    zmm028.to_excel(bases_dir / config.ZMM028_FILENAME, index=False)

    manual_path = bases_dir / "planilha_manual_quadro_diario.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    aba1 = wb.create_sheet("Indicadores Diarios")
    aba1.append(["Data", "Notas Aguardando Lancamento", "NF Pendente Faturamento", "Devolucao", "Scanner Documentos", "Intercompany"])
    aba1.append(["10/08/2026", 5, 3, 4, 0, 2])

    aba2 = wb.create_sheet("Pontos e Avisos")
    aba2.append(["PONTOS DE ATENÇÃO"])
    aba2.append(["Ponto de teste"])
    aba2.append(["AVISOS IMPORTANTES"])
    aba2.append(["Aviso de teste"])

    aba3 = wb.create_sheet("Datas Importantes")
    aba3.append(["Nome", "Periodo Inicio", "Periodo Fim", "Duracao (dias)", "Aprovacao", "Status"])
    aba3.append(["Fulano de Tal", "01/09/2026", "10/09/2026", 10, "OK", "Programada"])

    wb.create_sheet("Recebimento do Dia")

    wb.save(manual_path)

    return str(bases_dir), str(manual_path)


def _preparar_index(tmp_path):
    index_path = tmp_path / "index.html"
    index_path.write_text(INDEX_HTML_SINTETICO, encoding="utf-8")
    return str(index_path)


@pytest.fixture(autouse=True)
def _isolar_repo_root(tmp_path, monkeypatch):
    """congelar_dia grava os 4 JSON na raiz do repo real por padrão (REPO_ROOT do
    módulo) — redireciona pro tmp_path do teste, senão os testes escreveriam em cima
    dos históricos reais do projeto."""
    monkeypatch.setattr(cabecalho, "REPO_ROOT", str(tmp_path))
    return tmp_path


def test_checar_arquivos_lista_o_que_falta(tmp_path):
    faltando = cabecalho._checar_arquivos(str(tmp_path), str(tmp_path / "manual.xlsx"))
    assert len(faltando) == 4


def test_executar_aborta_sem_alterar_nada_quando_usuario_recusa_congelar(tmp_path, monkeypatch):
    bases_dir, manual_path = _preparar_bases(tmp_path)
    index_path = _preparar_index(tmp_path)
    monkeypatch.setattr("builtins.input", lambda _: "n")

    codigo = cabecalho.executar(bases_dir=bases_dir, manual_path=manual_path, index_path=index_path)

    assert codigo == 0
    assert not (tmp_path / "historico_mb51.json").exists()


def test_executar_congela_mas_nao_sobe_quando_usuario_recusa_push(tmp_path, monkeypatch):
    bases_dir, manual_path = _preparar_bases(tmp_path)
    index_path = _preparar_index(tmp_path)
    respostas = iter(["s", "n"])
    monkeypatch.setattr("builtins.input", lambda _: next(respostas))

    codigo = cabecalho.executar(bases_dir=bases_dir, manual_path=manual_path, index_path=index_path)

    assert codigo == 0
    historico = json.loads((tmp_path / "historico_mb51.json").read_text(encoding="utf-8"))
    assert "2026-08-10" in historico
    assert historico["2026-08-10"]["intercompany"] == 2


def test_executar_erro_quando_planilha_manual_nao_tem_linha_para_hoje(tmp_path, monkeypatch):
    bases_dir, manual_path = _preparar_bases(tmp_path)
    index_path = _preparar_index(tmp_path)

    # sobrescreve a planilha manual só com o dia 09/08 -> não bate com "hoje" (10/08,
    # detectado pela data mais recente da MB51)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    aba1 = wb.create_sheet("Indicadores Diarios")
    aba1.append(["Data", "Notas Aguardando Lancamento", "NF Pendente Faturamento", "Devolucao", "Scanner Documentos", "Intercompany"])
    aba1.append(["09/08/2026", 1, 1, 1, 1, 1])
    wb.create_sheet("Pontos e Avisos")
    aba3 = wb.create_sheet("Datas Importantes")
    aba3.append(["Nome", "Periodo Inicio", "Periodo Fim", "Duracao (dias)", "Aprovacao", "Status"])
    wb.create_sheet("Recebimento do Dia")
    wb.save(manual_path)

    monkeypatch.setattr("builtins.input", lambda _: pytest.fail("não deveria pedir confirmação"))
    codigo = cabecalho.executar(bases_dir=bases_dir, manual_path=manual_path, index_path=index_path)

    assert codigo == 1
    assert not (tmp_path / "historico_mb51.json").exists()


def test_executar_recusa_congelar_dia_ja_congelado(tmp_path, monkeypatch):
    bases_dir, manual_path = _preparar_bases(tmp_path)
    index_path = _preparar_index(tmp_path)
    (tmp_path / "historico_mb51.json").write_text(json.dumps({"2026-08-10": {"linhas_atendidas_d009": 0}}), encoding="utf-8")

    # "n" pra "Quer forçar o reprocessamento?" — pergunta antes de qualquer resumo,
    # então nem chega a perguntar "Posso congelar?" nem a pedir senha.
    monkeypatch.setattr("builtins.input", lambda _: "n")
    codigo = cabecalho.executar(bases_dir=bases_dir, manual_path=manual_path, index_path=index_path)

    assert codigo == 0
    # não deve ter sobrescrito o JSON existente
    historico = json.loads((tmp_path / "historico_mb51.json").read_text(encoding="utf-8"))
    assert historico == {"2026-08-10": {"linhas_atendidas_d009": 0}}


def test_executar_recusa_forcar_com_senha_errada(tmp_path, monkeypatch):
    bases_dir, manual_path = _preparar_bases(tmp_path)
    index_path = _preparar_index(tmp_path)
    (tmp_path / "historico_mb51.json").write_text(json.dumps({"2026-08-10": {"linhas_atendidas_d009": 0}}), encoding="utf-8")

    # "s" pra "Quer forçar?" — senha errada barra o resto antes de qualquer resumo
    # ser mostrado (não chega a perguntar "Posso congelar?").
    monkeypatch.setattr("builtins.input", lambda _: "s")
    monkeypatch.setattr(cabecalho.getpass, "getpass", lambda _: "senha-errada")
    codigo = cabecalho.executar(bases_dir=bases_dir, manual_path=manual_path, index_path=index_path)

    assert codigo == 0
    historico = json.loads((tmp_path / "historico_mb51.json").read_text(encoding="utf-8"))
    assert historico == {"2026-08-10": {"linhas_atendidas_d009": 0}}


def test_executar_forca_reprocessamento_com_senha_correta(tmp_path, monkeypatch):
    bases_dir, manual_path = _preparar_bases(tmp_path)
    index_path = _preparar_index(tmp_path)
    (tmp_path / "historico_mb51.json").write_text(json.dumps({"2026-08-10": {"linhas_atendidas_d009": 0}}), encoding="utf-8")

    # senha de teste — nunca a senha real de produção (esse arquivo vai pro GitHub).
    senha_teste = "senha-de-teste-123"
    monkeypatch.setattr(config, "SENHA_FORCAR_RECONGELAMENTO_SHA256", hashlib.sha256(senha_teste.encode("utf-8")).hexdigest())

    # "s" pra "Quer forçar?", "s" pra "Posso congelar?", "n" pra "Posso subir pro GitHub?"
    respostas = iter(["s", "s", "n"])
    monkeypatch.setattr("builtins.input", lambda _: next(respostas))
    monkeypatch.setattr(cabecalho.getpass, "getpass", lambda _: senha_teste)
    codigo = cabecalho.executar(bases_dir=bases_dir, manual_path=manual_path, index_path=index_path)

    assert codigo == 0
    historico = json.loads((tmp_path / "historico_mb51.json").read_text(encoding="utf-8"))
    assert historico["2026-08-10"]["intercompany"] == 2
    assert historico["2026-08-10"] != {"linhas_atendidas_d009": 0}


# ---- sincronização automática com o GitHub, no início de executar() -------

def _preparar_repo_sincronizado(tmp_path):
    """Inicializa tmp_path (== REPO_ROOT do teste, via _isolar_repo_root) como repo git
    com origin apontando pra um bare remoto, já sincronizado (1 commit inicial
    enviado). Retorna o path do bare remoto."""
    remoto = tmp_path.parent / f"{tmp_path.name}_remoto.git"
    subprocess.run(["git", "init", "--bare", str(remoto)], check=True, capture_output=True)

    subprocess.run(["git", "init"], cwd=tmp_path, check=True, capture_output=True)
    subprocess.run(["git", "config", "user.email", "teste@example.com"], cwd=tmp_path, check=True, capture_output=True)
    subprocess.run(["git", "config", "user.name", "Teste"], cwd=tmp_path, check=True, capture_output=True)
    subprocess.run(["git", "remote", "add", "origin", str(remoto)], cwd=tmp_path, check=True, capture_output=True)
    (tmp_path / "README.txt").write_text("inicial", encoding="utf-8")
    subprocess.run(["git", "add", "-A"], cwd=tmp_path, check=True, capture_output=True)
    subprocess.run(["git", "commit", "-m", "inicial"], cwd=tmp_path, check=True, capture_output=True)
    branch = subprocess.run(["git", "branch", "--show-current"], cwd=tmp_path, capture_output=True, text=True, check=True).stdout.strip()
    subprocess.run(["git", "push", "-u", "origin", branch], cwd=tmp_path, check=True, capture_output=True)
    return remoto


def _empurrar_commit_de_outro_computador(remoto, tmp_path, nome_arquivo, conteudo, mensagem):
    """Simula outra máquina clonando o mesmo remoto, commitando e enviando — sem
    passar pelo clone local em tmp_path, pra reproduzir "outra pessoa congelou em
    outro computador sem este saber"."""
    outro = tmp_path.parent / f"{tmp_path.name}_outro_pc"
    subprocess.run(["git", "clone", str(remoto), str(outro)], check=True, capture_output=True)
    subprocess.run(["git", "config", "user.email", "teste@example.com"], cwd=outro, check=True, capture_output=True)
    subprocess.run(["git", "config", "user.name", "Teste"], cwd=outro, check=True, capture_output=True)
    (outro / nome_arquivo).write_text(conteudo, encoding="utf-8")
    subprocess.run(["git", "add", "-A"], cwd=outro, check=True, capture_output=True)
    subprocess.run(["git", "commit", "-m", mensagem], cwd=outro, check=True, capture_output=True)
    subprocess.run(["git", "push"], cwd=outro, check=True, capture_output=True)


def test_executar_atualiza_sozinho_quando_esta_atras_do_remoto(tmp_path, monkeypatch, capsys):
    """Outro computador já congelou e enviou um dia; este computador nunca deu
    git pull. executar() tem que se atualizar sozinho (fast-forward), sem pedir
    nada pra ninguém, senão a checagem local de "já congelado?" ficaria cega pro
    que já existe no GitHub."""
    remoto = _preparar_repo_sincronizado(tmp_path)
    bases_dir, manual_path = _preparar_bases(tmp_path)
    index_path = _preparar_index(tmp_path)

    _empurrar_commit_de_outro_computador(remoto, tmp_path, "outro.txt", "outro dia", "Congela dia 2026-08-09")

    monkeypatch.setattr("builtins.input", lambda _: "n")  # recusa "Posso congelar?"
    codigo = cabecalho.executar(bases_dir=bases_dir, manual_path=manual_path, index_path=index_path)

    saida = capsys.readouterr().out
    assert codigo == 0
    assert "Repositório atualizado automaticamente com 1 commit(s)" in saida
    assert (tmp_path / "outro.txt").exists()


def test_executar_avisa_mas_nao_bloqueia_quando_esta_a_frente_do_remoto(tmp_path, monkeypatch, capsys):
    """Este computador tem um commit local pendente de envio (ex.: "Congela dia" de
    uma execução anterior) mas o remoto não mudou — não é motivo pra bloquear, só
    pra avisar; o push acontece normalmente ao final se o usuário confirmar."""
    _preparar_repo_sincronizado(tmp_path)

    (tmp_path / "pendente.txt").write_text("commit nao enviado", encoding="utf-8")
    subprocess.run(["git", "add", "-A"], cwd=tmp_path, check=True, capture_output=True)
    subprocess.run(["git", "commit", "-m", "Congela dia 2026-08-13"], cwd=tmp_path, check=True, capture_output=True)

    bases_dir, manual_path = _preparar_bases(tmp_path)
    index_path = _preparar_index(tmp_path)

    monkeypatch.setattr("builtins.input", lambda _: "n")  # recusa "Posso congelar?"
    codigo = cabecalho.executar(bases_dir=bases_dir, manual_path=manual_path, index_path=index_path)

    saida = capsys.readouterr().out
    assert codigo == 0
    assert "Aviso: este computador tem 1 commit(s) local(is) ainda não enviado(s)" in saida

    log_local = subprocess.run(
        ["git", "log", "-1", "--format=%s"], cwd=tmp_path, capture_output=True, text=True, check=True
    ).stdout
    assert "Congela dia 2026-08-13" in log_local


def test_executar_bloqueia_quando_diverge_do_remoto(tmp_path, monkeypatch, capsys):
    """Outra pessoa congelou e enviou um dia em outro computador, ENQUANTO este
    computador também tinha um commit local pendente de envio — histórico
    divergiu dos dois lados. index.html tem linhas de até 150 mil caracteres que
    git não mescla automaticamente, então isso tem que bloquear e pedir
    intervenção manual, nunca tentar resolver sozinho."""
    remoto = _preparar_repo_sincronizado(tmp_path)

    _empurrar_commit_de_outro_computador(remoto, tmp_path, "outro.txt", "outro dia", "Congela dia 2026-08-13")

    (tmp_path / "local.txt").write_text("dia daqui", encoding="utf-8")
    subprocess.run(["git", "add", "-A"], cwd=tmp_path, check=True, capture_output=True)
    subprocess.run(["git", "commit", "-m", "Congela dia 2026-08-14"], cwd=tmp_path, check=True, capture_output=True)

    monkeypatch.setattr("builtins.input", lambda _: pytest.fail("não deveria pedir nada — tem que bloquear antes"))
    codigo = cabecalho.executar()

    assert codigo == 1
    assert "divergiu" in capsys.readouterr().out


def test_executar_bloqueia_quando_fetch_falha(tmp_path, monkeypatch, capsys):
    subprocess.run(["git", "init"], cwd=tmp_path, check=True, capture_output=True)
    subprocess.run(["git", "config", "user.email", "teste@example.com"], cwd=tmp_path, check=True, capture_output=True)
    subprocess.run(["git", "config", "user.name", "Teste"], cwd=tmp_path, check=True, capture_output=True)
    subprocess.run(["git", "remote", "add", "origin", str(tmp_path / "remoto-inexistente")], cwd=tmp_path, check=True, capture_output=True)
    (tmp_path / "arquivo.txt").write_text("x", encoding="utf-8")
    subprocess.run(["git", "add", "-A"], cwd=tmp_path, check=True, capture_output=True)
    subprocess.run(["git", "commit", "-m", "inicial"], cwd=tmp_path, check=True, capture_output=True)

    monkeypatch.setattr("builtins.input", lambda _: pytest.fail("não deveria pedir nada — tem que bloquear antes"))
    codigo = cabecalho.executar()

    assert codigo == 1
    assert "git fetch falhou" in capsys.readouterr().out


# ---- git ------------------------------------------------------------------

def test_subir_para_github_faz_add_commit_push(tmp_path):
    remoto = tmp_path / "remoto.git"
    subprocess.run(["git", "init", "--bare", str(remoto)], check=True, capture_output=True)

    repo = tmp_path / "repo"
    repo.mkdir()
    subprocess.run(["git", "init"], cwd=repo, check=True, capture_output=True)
    subprocess.run(["git", "config", "user.email", "teste@example.com"], cwd=repo, check=True, capture_output=True)
    subprocess.run(["git", "config", "user.name", "Teste"], cwd=repo, check=True, capture_output=True)
    subprocess.run(["git", "remote", "add", "origin", str(remoto)], cwd=repo, check=True, capture_output=True)
    (repo / "arquivo.txt").write_text("inicial", encoding="utf-8")
    subprocess.run(["git", "add", "-A"], cwd=repo, check=True, capture_output=True)
    subprocess.run(["git", "commit", "-m", "inicial"], cwd=repo, check=True, capture_output=True)
    branch = subprocess.run(["git", "branch", "--show-current"], cwd=repo, capture_output=True, text=True, check=True).stdout.strip()
    subprocess.run(["git", "push", "-u", "origin", branch], cwd=repo, check=True, capture_output=True)

    (repo / "arquivo.txt").write_text("modificado", encoding="utf-8")
    cabecalho.subir_para_github(str(repo), "2026-08-10")

    log = subprocess.run(["git", "log", "-1", "--format=%s"], cwd=repo, capture_output=True, text=True, check=True)
    assert "Congela dia 2026-08-10" in log.stdout


def test_subir_para_github_envia_commit_pendente_mesmo_sem_mudanca_nova(tmp_path, capsys):
    """Reproduz o caso relatado por Fernando e Luiz: um "Congela dia ..." já commitado
    localmente (ex.: push anterior pulado ou que falhou) tem que ser enviado na
    próxima execução, mesmo que a working tree já esteja limpa e não haja nada novo
    pra commitar hoje."""
    remoto = tmp_path / "remoto.git"
    subprocess.run(["git", "init", "--bare", str(remoto)], check=True, capture_output=True)

    repo = tmp_path / "repo"
    repo.mkdir()
    subprocess.run(["git", "init"], cwd=repo, check=True, capture_output=True)
    subprocess.run(["git", "config", "user.email", "teste@example.com"], cwd=repo, check=True, capture_output=True)
    subprocess.run(["git", "config", "user.name", "Teste"], cwd=repo, check=True, capture_output=True)
    subprocess.run(["git", "remote", "add", "origin", str(remoto)], cwd=repo, check=True, capture_output=True)
    (repo / "arquivo.txt").write_text("inicial", encoding="utf-8")
    subprocess.run(["git", "add", "-A"], cwd=repo, check=True, capture_output=True)
    subprocess.run(["git", "commit", "-m", "inicial"], cwd=repo, check=True, capture_output=True)
    branch = subprocess.run(["git", "branch", "--show-current"], cwd=repo, capture_output=True, text=True, check=True).stdout.strip()
    subprocess.run(["git", "push", "-u", "origin", branch], cwd=repo, check=True, capture_output=True)

    # commit local que nunca chegou a ser enviado ao remoto
    (repo / "arquivo.txt").write_text("congelado ontem", encoding="utf-8")
    subprocess.run(["git", "add", "-A"], cwd=repo, check=True, capture_output=True)
    subprocess.run(["git", "commit", "-m", "Congela dia 2026-08-13"], cwd=repo, check=True, capture_output=True)

    # working tree já está limpa de novo — nada novo pra congelar hoje
    cabecalho.subir_para_github(str(repo), "2026-08-14")

    assert "Push concluído." in capsys.readouterr().out
    log_remoto = subprocess.run(
        ["git", "log", "-1", "--format=%s"], cwd=remoto, capture_output=True, text=True, check=True
    )
    assert "Congela dia 2026-08-13" in log_remoto.stdout


def test_subir_para_github_nao_faz_nada_se_nao_ha_mudanca(tmp_path, capsys):
    repo = tmp_path / "repo"
    repo.mkdir()
    subprocess.run(["git", "init"], cwd=repo, check=True, capture_output=True)
    subprocess.run(["git", "config", "user.email", "teste@example.com"], cwd=repo, check=True, capture_output=True)
    subprocess.run(["git", "config", "user.name", "Teste"], cwd=repo, check=True, capture_output=True)
    (repo / "arquivo.txt").write_text("inicial", encoding="utf-8")
    subprocess.run(["git", "add", "-A"], cwd=repo, check=True, capture_output=True)
    subprocess.run(["git", "commit", "-m", "inicial"], cwd=repo, check=True, capture_output=True)

    cabecalho.subir_para_github(str(repo), "2026-08-10")
    assert "Nada para commitar" in capsys.readouterr().out
