import datetime

import openpyxl
import pytest

from backend import planilha_manual


def _escrever_planilha(caminho, indicadores_linhas=None, pontos_avisos_linhas=None, datas_importantes_linhas=None):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    aba1 = wb.create_sheet(planilha_manual.ABA_INDICADORES)
    aba1.append(["Quadro Diário PDL – Indicadores Manuais"])
    aba1.append(["Preencher uma linha por dia."])
    aba1.append([])
    aba1.append(["Data", "Notas Aguardando Lancamento", "NF Pendente Faturamento", "Devolucao", "Scanner Documentos", "Intercompany"])
    for linha in indicadores_linhas or []:
        aba1.append(linha)

    aba2 = wb.create_sheet(planilha_manual.ABA_PONTOS_AVISOS)
    for linha in pontos_avisos_linhas or []:
        aba2.append([linha])

    aba3 = wb.create_sheet(planilha_manual.ABA_DATAS_IMPORTANTES)
    aba3.append(["Matriz de Férias – Almoxarifado PDL"])
    aba3.append(["Status: Quitada / Programada / Aguardando aprovação"])
    aba3.append([])
    aba3.append(["Nome", "Periodo Inicio", "Periodo Fim", "Duracao (dias)", "Aprovacao", "Status"])
    for linha in datas_importantes_linhas or []:
        aba3.append(linha)

    aba4 = wb.create_sheet("Recebimento do Dia")
    aba4.append(["PROPOSITALMENTE VAZIO"])

    wb.save(caminho)
    return str(caminho)


# ---- Indicadores Diarios ---------------------------------------------------

def test_le_uma_linha_por_dia_com_as_5_chaves(tmp_path):
    caminho = _escrever_planilha(
        tmp_path / "manual.xlsx",
        indicadores_linhas=[
            ["10/08/2026", 5, 3, 4, 0, 2],
            ["11/08/2026", 1, 0, 0, 1, 0],
        ],
    )
    resultado = planilha_manual.ler_indicadores_diarios(caminho)
    assert set(resultado.keys()) == {"2026-08-10", "2026-08-11"}
    assert resultado["2026-08-10"] == {
        "notas_aguardando_lancamento": 5,
        "nf_pendente_faturamento": 3,
        "devolucao": 4,
        "scanner_documentos": 0,
        "intercompany": 2,
    }
    assert resultado["2026-08-11"]["intercompany"] == 0


def test_planilha_vazia_retorna_dict_vazio(tmp_path):
    caminho = _escrever_planilha(tmp_path / "manual.xlsx", indicadores_linhas=[])
    assert planilha_manual.ler_indicadores_diarios(caminho) == {}


def test_data_duplicada_lanca_erro(tmp_path):
    caminho = _escrever_planilha(
        tmp_path / "manual.xlsx",
        indicadores_linhas=[["10/08/2026", 1, 1, 1, 1, 1], ["10/08/2026", 2, 2, 2, 2, 2]],
    )
    with pytest.raises(ValueError, match="Datas duplicadas"):
        planilha_manual.ler_indicadores_diarios(caminho)


def test_celula_vazia_lanca_erro_em_vez_de_assumir_zero(tmp_path):
    caminho = _escrever_planilha(
        tmp_path / "manual.xlsx",
        indicadores_linhas=[["10/08/2026", None, 3, 4, 0, 2]],
    )
    with pytest.raises(ValueError, match="vazia"):
        planilha_manual.ler_indicadores_diarios(caminho)


# ---- Pontos e Avisos --------------------------------------------------------

def test_le_pontos_e_avisos_separados_por_secao(tmp_path):
    caminho = _escrever_planilha(
        tmp_path / "manual.xlsx",
        pontos_avisos_linhas=[
            "Pontos de Atenção e Avisos Importantes",
            "Cada linha vira um item de bullet no painel.",
            None,
            "PONTOS DE ATENÇÃO",
            "Estorno de materiais para o estoque, somente com chamado aprovado.",
            "Período fechado para entrada NFs",
            None,
            None,
            "AVISOS IMPORTANTES",
            "Todos os estornos devem ser feitos na data de hoje, não retroativa.",
        ],
    )
    resultado = planilha_manual.ler_pontos_avisos(caminho)
    assert resultado["pontos_atencao"] == [
        "Estorno de materiais para o estoque, somente com chamado aprovado.",
        "Período fechado para entrada NFs",
    ]
    assert resultado["avisos_importantes"] == ["Todos os estornos devem ser feitos na data de hoje, não retroativa."]


def test_linha_em_branco_nao_fecha_secao(tmp_path):
    caminho = _escrever_planilha(
        tmp_path / "manual.xlsx",
        pontos_avisos_linhas=["PONTOS DE ATENÇÃO", "Item 1", None, "Item 2"],
    )
    resultado = planilha_manual.ler_pontos_avisos(caminho)
    assert resultado["pontos_atencao"] == ["Item 1", "Item 2"]


def test_sem_secoes_retorna_listas_vazias(tmp_path):
    caminho = _escrever_planilha(tmp_path / "manual.xlsx", pontos_avisos_linhas=["Só um título qualquer"])
    resultado = planilha_manual.ler_pontos_avisos(caminho)
    assert resultado == {"pontos_atencao": [], "avisos_importantes": []}


# ---- Datas Importantes -------------------------------------------------------

def test_le_matriz_de_ferias(tmp_path):
    caminho = _escrever_planilha(
        tmp_path / "manual.xlsx",
        datas_importantes_linhas=[
            ["Luiz Sergio", "23/03/2026", "05/04/2026", 14, "OK", "Quitada"],
            ["Fernando Luis", "17/08/2026", "05/09/2026", None, "OK", "Programada"],
        ],
    )
    resultado = planilha_manual.ler_datas_importantes(caminho)
    assert len(resultado) == 2
    assert resultado[0] == {
        "nome": "Luiz Sergio",
        "periodo_inicio": datetime.date(2026, 3, 23),
        "periodo_fim": datetime.date(2026, 4, 5),
        "duracao_dias": 14,
        "aprovacao": "OK",
        "status": "Quitada",
    }
    assert resultado[1]["duracao_dias"] is None


def test_periodo_invalido_lanca_erro(tmp_path):
    caminho = _escrever_planilha(
        tmp_path / "manual.xlsx",
        datas_importantes_linhas=[["Fulano", "não é data", "05/04/2026", 14, "OK", "Quitada"]],
    )
    with pytest.raises(ValueError, match="Período inválido"):
        planilha_manual.ler_datas_importantes(caminho)


# ---- normalização de acento/caixa -------------------------------------------

def test_sem_acento_maiusculo():
    assert planilha_manual.sem_acento_maiusculo("Aguardando aprovação") == "AGUARDANDO APROVACAO"
    assert planilha_manual.sem_acento_maiusculo("  Atenção  ") == "ATENCAO"
